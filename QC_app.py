# QC_app.py
# -*- coding: utf-8 -*-
import streamlit as st
import pandas as pd
import numpy as np
import io
import re
import os
from datetime import datetime
from copy import copy
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from database import init_db, save_record, load_all_records, delete_records

# ==================== 项目规则配置 ====================
PROJECT_CONFIGS = {
    "新冠甲乙流": {
        "channels": ["CY5", "FAM", "Texas Red", "VIC"],
        "channel_labels": {
            "CY5": "CY5通道Ct值（内标）",
            "FAM": "FAM通道Ct值（甲流）",
            "Texas Red": "Texas Red通道Ct值（乙流）",
            "VIC": "VIC通道Ct值（新冠）"
        },
        "pathogens": [
            {"name": "甲型流感病毒", "channel": "FAM", "threshold": 38},
            {"name": "乙型流感病毒", "channel": "Texas Red", "threshold": 38},
            {"name": "2019-nCoV新型冠状病毒", "channel": "VIC", "threshold": 38},
        ],
        "use_prefix": True,
        "reference_categories": {
            "N": "阴性参考品",
            "P": "阳性参考品",
            "S": "最低检出限参考品",
            "R1": "重复性参考品R1",
            "R2": "重复性参考品R2",
            "R3": "重复性参考品R3",
            "YANG": "ABnC阳性质控品",
            "YIN": "ABnC阴性质控品"
        },
        "judge_rules": {
            "N": {
                "CY5": "≤38", "FAM": "Undetermined或≥42", "Texas Red": "Undetermined或≥42", "VIC": "Undetermined或≥42",
                "expected": "阴性", "quality": "均为阴性",
                "rule_text": "\"FAM通道Ct值\"为\"Undetermined或Ct≥42\"\n\"Texas Red通道Ct值\"为\"Undetermined或Ct≥42\"\n\"VIC通道Ct值\"为\"Undetermined或Ct≥42\"\n\"CY5通道Ct值\"为\"Ct≤38\""
            },
            "P1-P10": {
                "CY5": "≤38", "FAM": "≤38", "Texas Red": "Undetermined或≥42", "VIC": "Undetermined或≥42",
                "expected": "阳性", "quality": "甲型流感病毒阳性",
                "rule_text": "\"FAM通道Ct值\"为\"Ct≤38\"\n\"Texas Red通道Ct值\"为\"Undetermined或Ct≥42\"\n\"VIC通道Ct值\"为\"Undetermined或Ct≥42\"\n\"CY5通道Ct值\"为\"Ct≤38\""
            },
            "P11-P14": {
                "CY5": "≤38", "FAM": "Undetermined或≥42", "Texas Red": "≤38", "VIC": "Undetermined或≥42",
                "expected": "阳性", "quality": "乙型流感病毒阳性",
                "rule_text": "\"FAM通道Ct值\"为\"Undetermined或Ct≥42\"\n\"Texas Red通道Ct值\"为\"Ct≤38\"\n\"VIC通道Ct值\"为\"Undetermined或Ct≥42\"\n\"CY5通道Ct值\"为\"Ct≤38\""
            },
            "P15-P20": {
                "CY5": "≤38", "FAM": "Undetermined或≥42", "Texas Red": "Undetermined或≥42", "VIC": "≤38",
                "expected": "阳性", "quality": "2019-nCoV新型冠状病毒阳性",
                "rule_text": "\"FAM通道Ct值\"为\"Undetermined或Ct≥42\"\n\"Texas Red通道Ct值\"为\"Undetermined或Ct≥42\"\n\"VIC通道Ct值\"为\"Ct≤38\"\n\"CY5通道Ct值\"为\"Ct≤38\""
            },
            "S1-S5": {
                "CY5": "≤38", "FAM": "≤38", "Texas Red": "Undetermined或≥42", "VIC": "Undetermined或≥42",
                "expected": "阳性", "quality": "甲型流感病毒阳性",
                "rule_text": "\"FAM通道Ct值\"为\"Ct≤38\"\n\"Texas Red通道Ct值\"为\"Undetermined或Ct≥42\"\n\"VIC通道Ct值\"为\"Undetermined或Ct≥42\"\n\"CY5通道Ct值\"为\"Ct≤38\""
            },
            "S6-S7": {
                "CY5": "≤38", "FAM": "Undetermined或≥42", "Texas Red": "≤38", "VIC": "Undetermined或≥42",
                "expected": "阳性", "quality": "乙型流感病毒阳性",
                "rule_text": "\"FAM通道Ct值\"为\"Undetermined或Ct≥42\"\n\"Texas Red通道Ct值\"为\"Ct≤38\"\n\"VIC通道Ct值\"为\"Undetermined或Ct≥42\"\n\"CY5通道Ct值\"为\"Ct≤38\""
            },
            "S8": {
                "CY5": "≤38", "FAM": "Undetermined或≥42", "Texas Red": "Undetermined或≥42", "VIC": "≤38",
                "expected": "阳性", "quality": "2019-nCoV新型冠状病毒阳性",
                "rule_text": "\"FAM通道Ct值\"为\"Undetermined或Ct≥42\"\n\"Texas Red通道Ct值\"为\"Undetermined或Ct≥42\"\n\"VIC通道Ct值\"为\"Ct≤38\"\n\"CY5通道Ct值\"为\"Ct≤38\""
            },
            "R1": {
                "CY5": "无要求", "FAM": "≤38", "Texas Red": "≤38", "VIC": "≤38",
                "expected": "阳性",
                "quality": "检测重复性参考品R1，重复检测10次，R1检测结果应均为甲型流感病毒阳性、乙型流感病毒阳性及新型冠状病毒阳性，且各重复性参考品检测结果Ct值的变异系数CV值均≤5%（内标通道无需进行统计）。",
                "rule_text": "\"FAM通道Ct值\"为\"Ct≤38\"\n\"Texas Red通道Ct值\"为\"Ct≤38\"\n\"VIC通道Ct值\"为\"Ct≤38\"\n\"CY5通道Ct值\"为\"Ct≤38\""
            },
            "R2": {
                "CY5": "无要求", "FAM": "≤38", "Texas Red": "≤38", "VIC": "≤38",
                "expected": "阳性",
                "quality": "检测重复性参考品R2，重复检测10次，R2检测结果应均为甲型流感病毒阳性、乙型流感病毒阳性及新型冠状病毒阳性，且各重复性参考品检测结果Ct值的变异系数CV值均≤5%（内标通道无需进行统计）。",
                "rule_text": "\"FAM通道Ct值\"为\"Ct≤38\"\n\"Texas Red通道Ct值\"为\"Ct≤38\"\n\"VIC通道Ct值\"为\"Ct≤38\"\n\"CY5通道Ct值\"为\"Ct≤38\""
            },
            "R3": {
                "CY5": "≤38", "FAM": "Undetermined或≥42", "Texas Red": "Undetermined或≥42", "VIC": "Undetermined或≥42",
                "expected": "阴性",
                "quality": "检测重复性参考品R3，重复检测10次，R3检测结果应为阴性。",
                "rule_text": "\"FAM通道Ct值\"为\"Undetermined或Ct≥42\"\n\"Texas Red通道Ct值\"为\"Undetermined或Ct≥42\"\n\"VIC通道Ct值\"为\"Undetermined或Ct≥42\"\n\"CY5通道Ct值\"为\"Ct≤38\""
            },
            "YANG": {
                "CY5": "无要求", "FAM": "≤38", "Texas Red": "≤38", "VIC": "≤38",
                "expected": "阳性",
                "quality": "FAM、Texas Red、VIC检测通道均存在明显扩增曲线，且Ct值≤32，CY5通道有或无扩增曲线",
                "rule_text": "\"FAM通道Ct值\"为\"Ct≤38\"\n\"Texas Red通道Ct值\"为\"Ct≤38\"\n\"VIC通道Ct值\"为\"Ct≤38\"\n\"CY5通道Ct值\"为\"Ct≤38\""
            },
            "YIN": {
                "CY5": "≤38", "FAM": "Undetermined", "Texas Red": "Undetermined", "VIC": "Undetermined",
                "expected": "阴性",
                "quality": "为阴性，CY5通道存在明显扩增曲线，且Ct值≤38，其他通道无扩增曲线。",
                "rule_text": "\"FAM通道Ct值\"为\"Undetermined\"\n\"Texas Red通道Ct值\"为\"Undetermined\"\n\"VIC通道Ct值\"为\"Undetermined\"\n\"CY5通道Ct值\"为\"Ct≤38\""
            },
        }
    },
    "通用": {
        "channels": ["CY5", "FAM", "Texas Red", "VIC"],
        "channel_labels": {
            "CY5": "CY5通道Ct值（内标）",
            "FAM": "FAM通道Ct值",
            "Texas Red": "Texas Red通道Ct值",
            "VIC": "VIC通道Ct值"
        },
        "pathogens": [
            {"name": "", "channel": "FAM", "threshold": 38},
            {"name": "", "channel": "Texas Red", "threshold": 38},
            {"name": "", "channel": "VIC", "threshold": 38},
        ],
        "use_prefix": False,
        "reference_categories": {
            "N": "阴性参考品",
            "P": "阳性参考品",
            "S": "最低检出限参考品",
            "R1": "重复性参考品R1",
            "R2": "重复性参考品R2",
            "R3": "重复性参考品R3",
            "YANG": "阳性质控品",
            "YIN": "阴性质控品"
        },
        "judge_rules": {
            "N": {
                "CY5": "≤38", "FAM": "Undetermined或≥42", "Texas Red": "Undetermined或≥42", "VIC": "Undetermined或≥42",
                "expected": "阴性", "quality": "均为阴性",
                "rule_text": "\"FAM通道Ct值\"为\"Undetermined或Ct≥42\"\n\"Texas Red通道Ct值\"为\"Undetermined或Ct≥42\"\n\"VIC通道Ct值\"为\"Undetermined或Ct≥42\"\n\"CY5通道Ct值\"为\"Ct≤38\""
            },
            "P1-P10": {
                "CY5": "≤38", "FAM": "≤38", "Texas Red": "Undetermined或≥42", "VIC": "Undetermined或≥42",
                "expected": "阳性", "quality": "阳性",
                "rule_text": "\"FAM通道Ct值\"为\"Ct≤38\"\n\"Texas Red通道Ct值\"为\"Undetermined或Ct≥42\"\n\"VIC通道Ct值\"为\"Undetermined或Ct≥42\"\n\"CY5通道Ct值\"为\"Ct≤38\""
            },
            "P11-P14": {
                "CY5": "≤38", "FAM": "Undetermined或≥42", "Texas Red": "≤38", "VIC": "Undetermined或≥42",
                "expected": "阳性", "quality": "阳性",
                "rule_text": "\"FAM通道Ct值\"为\"Undetermined或Ct≥42\"\n\"Texas Red通道Ct值\"为\"Ct≤38\"\n\"VIC通道Ct值\"为\"Undetermined或Ct≥42\"\n\"CY5通道Ct值\"为\"Ct≤38\""
            },
            "P15-P20": {
                "CY5": "≤38", "FAM": "Undetermined或≥42", "Texas Red": "Undetermined或≥42", "VIC": "≤38",
                "expected": "阳性", "quality": "阳性",
                "rule_text": "\"FAM通道Ct值\"为\"Undetermined或Ct≥42\"\n\"Texas Red通道Ct值\"为\"Undetermined或Ct≥42\"\n\"VIC通道Ct值\"为\"Ct≤38\"\n\"CY5通道Ct值\"为\"Ct≤38\""
            },
            "S1-S5": {
                "CY5": "≤38", "FAM": "≤38", "Texas Red": "Undetermined或≥42", "VIC": "Undetermined或≥42",
                "expected": "阳性", "quality": "阳性",
                "rule_text": "\"FAM通道Ct值\"为\"Ct≤38\"\n\"Texas Red通道Ct值\"为\"Undetermined或Ct≥42\"\n\"VIC通道Ct值\"为\"Undetermined或Ct≥42\"\n\"CY5通道Ct值\"为\"Ct≤38\""
            },
            "S6-S7": {
                "CY5": "≤38", "FAM": "Undetermined或≥42", "Texas Red": "≤38", "VIC": "Undetermined或≥42",
                "expected": "阳性", "quality": "阳性",
                "rule_text": "\"FAM通道Ct值\"为\"Undetermined或Ct≥42\"\n\"Texas Red通道Ct值\"为\"Ct≤38\"\n\"VIC通道Ct值\"为\"Undetermined或Ct≥42\"\n\"CY5通道Ct值\"为\"Ct≤38\""
            },
            "S8": {
                "CY5": "≤38", "FAM": "Undetermined或≥42", "Texas Red": "Undetermined或≥42", "VIC": "≤38",
                "expected": "阳性", "quality": "阳性",
                "rule_text": "\"FAM通道Ct值\"为\"Undetermined或Ct≥42\"\n\"Texas Red通道Ct值\"为\"Undetermined或Ct≥42\"\n\"VIC通道Ct值\"为\"Ct≤38\"\n\"CY5通道Ct值\"为\"Ct≤38\""
            },
            "R1": {
                "CY5": "无要求", "FAM": "≤38", "Texas Red": "≤38", "VIC": "≤38",
                "expected": "阳性",
                "quality": "检测重复性参考品R1，重复检测10次，R1检测结果应均为阳性，且各重复性参考品检测结果Ct值的变异系数CV值均≤5%（内标通道无需进行统计）。",
                "rule_text": "\"FAM通道Ct值\"为\"Ct≤38\"\n\"Texas Red通道Ct值\"为\"Ct≤38\"\n\"VIC通道Ct值\"为\"Ct≤38\"\n\"CY5通道Ct值\"为\"Ct≤38\""
            },
            "R2": {
                "CY5": "无要求", "FAM": "≤38", "Texas Red": "≤38", "VIC": "≤38",
                "expected": "阳性",
                "quality": "检测重复性参考品R2，重复检测10次，R2检测结果应均为阳性，且各重复性参考品检测结果Ct值的变异系数CV值均≤5%（内标通道无需进行统计）。",
                "rule_text": "\"FAM通道Ct值\"为\"Ct≤38\"\n\"Texas Red通道Ct值\"为\"Ct≤38\"\n\"VIC通道Ct值\"为\"Ct≤38\"\n\"CY5通道Ct值\"为\"Ct≤38\""
            },
            "R3": {
                "CY5": "≤38", "FAM": "Undetermined或≥42", "Texas Red": "Undetermined或≥42", "VIC": "Undetermined或≥42",
                "expected": "阴性",
                "quality": "检测重复性参考品R3，重复检测10次，R3检测结果应为阴性。",
                "rule_text": "\"FAM通道Ct值\"为\"Undetermined或Ct≥42\"\n\"Texas Red通道Ct值\"为\"Undetermined或Ct≥42\"\n\"VIC通道Ct值\"为\"Undetermined或Ct≥42\"\n\"CY5通道Ct值\"为\"Ct≤38\""
            },
            "YANG": {
                "CY5": "无要求", "FAM": "≤38", "Texas Red": "≤38", "VIC": "≤38",
                "expected": "阳性",
                "quality": "FAM、Texas Red、VIC检测通道均存在明显扩增曲线，且Ct值≤38，CY5通道有或无扩增曲线",
                "rule_text": "\"FAM通道Ct值\"为\"Ct≤38\"\n\"Texas Red通道Ct值\"为\"Ct≤38\"\n\"VIC通道Ct值\"为\"Ct≤38\"\n\"CY5通道Ct值\"为\"Ct≤38\""
            },
            "YIN": {
                "CY5": "≤38", "FAM": "Undetermined", "Texas Red": "Undetermined", "VIC": "Undetermined",
                "expected": "阴性",
                "quality": "为阴性，CY5通道存在明显扩增曲线，且Ct值≤38，其他通道无扩增曲线。",
                "rule_text": "\"FAM通道Ct值\"为\"Undetermined\"\n\"Texas Red通道Ct值\"为\"Undetermined\"\n\"VIC通道Ct值\"为\"Undetermined\"\n\"CY5通道Ct值\"为\"Ct≤38\""
            },
        }
    }
}

# ==================== 页面设置 ====================
st.set_page_config(page_title="QC数据智能分析系统", layout="wide")
st.title("QC数据智能分析系统")

init_db()

project_name = st.selectbox("选择项目", list(PROJECT_CONFIGS.keys()))
config = PROJECT_CONFIGS[project_name]

# ==================== 函数 ====================
def parse_range(prefix_str):
    result = []
    parts = prefix_str.split(",")
    for part in parts:
        part = part.strip()
        if "-" in part:
            match = re.match(r"([A-Za-z]+)(\d+)-([A-Za-z]+)?(\d+)", part)
            if match:
                prefix1 = match.group(1)
                start = int(match.group(2))
                prefix2 = match.group(3) if match.group(3) else prefix1
                end = int(match.group(4))
                for i in range(start, end + 1):
                    result.append(f"{prefix1}{i}")
        else:
            result.append(part)
    return result

def match_judge_rule(sample_name):
    judge_rules = config.get("judge_rules", {})
    s = str(sample_name)
    for key, rule in judge_rules.items():
        if re.match(r"^[A-Za-z]+\d*$", key):
            if s.startswith(key):
                return rule
        else:
            prefixes = parse_range(key)
            if s in prefixes:
                return rule
    return None

def check_channel(value, rule_str):
    if not rule_str or rule_str.strip() == "" or rule_str.strip() == "无要求":
        return True
    if value == "Undetermined" or value is None or (isinstance(value, float) and np.isnan(value)):
        return "Undetermined" in rule_str
    try:
        ct_val = float(value)
    except (ValueError, TypeError):
        return False
    if "Undetermined" in rule_str:
        return True
    nums = re.findall(r"[\d.]+", rule_str)
    if not nums:
        return False
    threshold = float(nums[0])
    if "≤" in rule_str:
        return ct_val <= threshold
    if "≥" in rule_str:
        return ct_val >= threshold
    return False

def do_judge(row_data, channels, judge_rule):
    cy5_val = row_data.get("CY5通道Ct值", "Undetermined")
    if cy5_val == "Undetermined" or (isinstance(cy5_val, (int, float)) and cy5_val > 38):
        return "无效", "不符合规定", "\"CY5通道Ct值\"为\"Undetermined或Ct>38\"，结果无效。"

    expected = judge_rule.get("expected", "") if judge_rule else ""
    pathogens = config["pathogens"]

    positive_count = 0
    for pathogen in pathogens:
        ch = pathogen["channel"]
        if ch not in channels:
            continue
        ch_val = row_data.get(f"{ch}通道Ct值", "Undetermined")
        try:
            ch_ct = float(ch_val)
            if ch_ct <= pathogen["threshold"]:
                positive_count += 1
        except (ValueError, TypeError):
            pass

    if positive_count >= 2 and expected == "阳性":
        rule_text = judge_rule.get("rule_text", "") if judge_rule else ""
        return "阳性", "符合规定", rule_text

    for pathogen in pathogens:
        ch = pathogen["channel"]
        if ch not in channels:
            continue
        ch_val = row_data.get(f"{ch}通道Ct值", "Undetermined")
        try:
            ch_ct = float(ch_val)
            if ch_ct <= pathogen["threshold"]:
                if config["use_prefix"] and pathogen["name"]:
                    result_name = f"{pathogen['name']}阳性"
                else:
                    result_name = "阳性"
                rule_text = judge_rule.get("rule_text", "") if judge_rule else ""
                return result_name, "符合规定", rule_text
        except (ValueError, TypeError):
            pass

    all_negative = True
    for pathogen in pathogens:
        ch = pathogen["channel"]
        if ch not in channels:
            continue
        ch_val = row_data.get(f"{ch}通道Ct值", "Undetermined")
        if ch_val == "Undetermined":
            continue
        try:
            ch_ct = float(ch_val)
            if ch_ct < 42:
                all_negative = False
                break
        except (ValueError, TypeError):
            pass

    if all_negative:
        rule_text = judge_rule.get("rule_text", "") if judge_rule else ""
        return "阴性", "符合规定", rule_text

    return "不符合", "不符合规定", ""

def get_category(sample_name):
    s = str(sample_name)
    cats = config.get("reference_categories", {})
    if re.match(r"^R1\d*$", s):
        return cats.get("R1", "重复性参考品R1")
    if re.match(r"^R2\d*$", s):
        return cats.get("R2", "重复性参考品R2")
    if re.match(r"^R3\d*$", s):
        return cats.get("R3", "重复性参考品R3")
    for prefix, cat in cats.items():
        if s.startswith(prefix) and prefix not in ["R1", "R2", "R3"]:
            return cat
    return ""

def get_quality(sample_name):
    s = str(sample_name)
    judge_rules = config.get("judge_rules", {})
    for key, rule in judge_rules.items():
        if re.match(r"^[A-Za-z]+\d*$", key):
            if s.startswith(key):
                return rule.get("quality", "")
        else:
            prefixes = parse_range(key)
            if s in prefixes:
                return rule.get("quality", "")
    return ""

def fmt_ct(val):
    if val == "Undetermined" or val is None or (isinstance(val, float) and np.isnan(val)):
        return "Undetermined"
    try:
        return float(val)
    except (ValueError, TypeError):
        return str(val)

# ==================== 侧边栏 ====================
st.sidebar.header("📋 基本信息")
product_name = st.sidebar.text_input("品名", value="")
batch_no = st.sidebar.text_input("批号", value="")
spec = st.sidebar.text_input("规格", value="")
inspector = st.sidebar.text_input("检验人", value="")
inspection_date = st.sidebar.date_input("检验日期", value=datetime.now().date())
ref_batch = st.sidebar.text_input("企业参考品批号", value="")

# ==================== 主区域 ====================
tab1, tab2 = st.tabs(["📤 上传数据 & 生成模板一", "📂 历史记录"])

with tab1:
    st.subheader("上传仪器原始数据")
    uploaded_file = st.file_uploader("选择仪器导出的 .xls 文件", type=["xls", "xlsx"])

    if uploaded_file is not None:
        try:
            df_raw = pd.read_excel(uploaded_file, header=6)
            if "Well" not in str(df_raw.columns) and "Sample Name" not in str(df_raw.columns):
                raise ValueError("列名不对")
        except:
            df_full = pd.read_excel(uploaded_file, header=None)
            header_row = None
            for i in range(len(df_full)):
                row_vals = df_full.iloc[i].astype(str).tolist()
                if "Well" in row_vals and "Sample Name" in row_vals:
                    header_row = i
                    break
            if header_row is None:
                st.error("找不到表头行，请确认文件格式。")
                st.stop()
            df_raw = pd.read_excel(uploaded_file, header=header_row)

        df_raw.columns = [str(c).strip() for c in df_raw.columns]
        df_raw = df_raw.dropna(how="all")

        ct_col = None
        for col in df_raw.columns:
            col_clean = col.replace(" ", "").upper()
            if col_clean in ["CT", "CТ"]:
                ct_col = col
                break
        if ct_col is None:
            for col in df_raw.columns:
                col_upper = col.replace(" ", "").upper()
                if "CT" in col_upper or "CТ" in col_upper:
                    ct_col = col
                    break
        if ct_col is None:
            st.error(f"找不到Ct值列，当前列名：{list(df_raw.columns)}")
            st.stop()

        st.subheader("📊 原始数据预览")
        st.dataframe(df_raw, use_container_width=True)

        available_targets = df_raw["Target Name"].dropna().unique().tolist()
        available_targets_clean = [t.strip().upper() for t in available_targets]
        channels = [t for t in config["channels"] if t.upper() in available_targets_clean]

        df_data = df_raw.dropna(subset=["Sample Name", "Target Name"], how="all").copy()
        df_data["Sample Name"] = df_data["Sample Name"].astype(str).str.strip()
        df_data["Target Name"] = df_data["Target Name"].astype(str).str.strip()
        df_data["Target_Clean"] = df_data["Target Name"].str.replace(" ", "").str.upper()

        samples = []
        non_r_seen = set()
        r_buffer = {}

        for idx, row in df_data.iterrows():
            sample = str(row["Sample Name"]).strip()
            if not sample:
                continue
            if sample.startswith("R"):
                if sample not in r_buffer:
                    r_buffer[sample] = {"count": 0}
                r_buffer[sample]["count"] += 1
                if r_buffer[sample]["count"] % 4 == 1:
                    samples.append(sample)
                if r_buffer[sample]["count"] % 4 == 0:
                    r_buffer[sample]["count"] = 0
            else:
                if sample not in non_r_seen:
                    samples.append(sample)
                    non_r_seen.add(sample)

        category_order = {"N": 1, "P": 2, "S": 3, "R": 4, "YANG": 5, "YIN": 6}
        def sort_key(sample):
            s = str(sample)
            for prefix, order in category_order.items():
                if s.startswith(prefix):
                    nums = re.findall(r"\d+", s)
                    num = int(nums[0]) if nums else 0
                    return (order, num)
            return (99, 0)

        samples = sorted(samples, key=sort_key)

        template_data = []
        current_category = ""
        r_occurrence = {}

        for sample in samples:
            category = get_category(str(sample))
            quality = get_quality(str(sample))
            judge_rule = match_judge_rule(str(sample))

            display_category = category if category != current_category else ""
            display_quality = quality if category != current_category else ""
            if category != current_category:
                current_category = category

            row_data = {
                "参考品": display_category,
                "编号": sample,
                "质量标准": display_quality,
            }

            if sample.startswith("R"):
                r_occurrence[sample] = r_occurrence.get(sample, 0) + 1
                occ = r_occurrence[sample]
                sample_all_rows = df_data[df_data["Sample Name"] == sample]
                group_start = (occ - 1) * 4
                group_end = occ * 4
                sample_rows = sample_all_rows.iloc[group_start:group_end]
            else:
                sample_rows = df_data[df_data["Sample Name"] == sample]

            for ch in channels:
                ch_clean = ch.replace(" ", "").upper()
                ch_row = sample_rows[sample_rows["Target_Clean"] == ch_clean]
                if len(ch_row) > 0:
                    ct_val = ch_row[ct_col].values[0]
                    row_data[f"{ch}通道Ct值"] = fmt_ct(ct_val)
                else:
                    row_data[f"{ch}通道Ct值"] = "Undetermined"

            if judge_rule:
                result, verdict, rule_text = do_judge(row_data, channels, judge_rule)
                row_data["检测结果"] = result
                row_data["结果判读"] = verdict
                if display_category != "":
                    row_data["结果判读规则"] = rule_text
                else:
                    row_data["结果判读规则"] = ""
            else:
                row_data["检测结果"] = ""
                row_data["结果判读"] = ""
                row_data["结果判读规则"] = ""

            template_data.append(row_data)

        # ==================== 在 R1/R2 数据后插入统计行 ====================
        r1_last_idx = -1
        r2_last_idx = -1
        for i, row in enumerate(template_data):
            if str(row["编号"]) == "R1":
                r1_last_idx = i
            elif str(row["编号"]) == "R2":
                r2_last_idx = i

        if r1_last_idx >= 0:
            r1_rows = [r for r in template_data if str(r["编号"]) == "R1"]
            if len(r1_rows) >= 2:
                avg_row = {"参考品": "", "编号": "平均值", "质量标准": ""}
                std_row = {"参考品": "", "编号": "标准偏差", "质量标准": ""}
                cv_row = {"参考品": "", "编号": "变异系数（CV值）", "质量标准": ""}
                cv_values = {}
                for ch in channels:
                    vals = []
                    for r in r1_rows:
                        v = r.get(f"{ch}通道Ct值", "Undetermined")
                        try:
                            vals.append(float(v))
                        except (ValueError, TypeError):
                            pass
                    if len(vals) > 0:
                        a = np.mean(vals)
                        s = np.std(vals, ddof=1) if len(vals) > 1 else 0.0
                        c = round(s / a * 100, 2) if a != 0 else 0
                    else:
                        a, s, c = "/", "/", "/"
                    avg_row[f"{ch}通道Ct值"] = a if not isinstance(a, str) else a
                    std_row[f"{ch}通道Ct值"] = s if not isinstance(s, str) else s
                    cv_row[f"{ch}通道Ct值"] = f"{c}%" if not isinstance(c, str) else c
                    cv_values[ch] = c
                avg_row["检测结果"] = "/"
                avg_row["结果判读"] = "/"
                avg_row["结果判读规则"] = ""
                std_row["检测结果"] = "/"
                std_row["结果判读"] = "/"
                std_row["结果判读规则"] = ""
                cv_row["检测结果"] = "/"
                cv_ok = all(isinstance(cv_values.get(ch), (int, float)) and cv_values.get(ch) <= 5 for ch in channels if ch != "CY5")
                cv_row["结果判读"] = "符合规定" if cv_ok else ""
                cv_row["结果判读规则"] = '数值小于等于"5"'
                template_data.insert(r1_last_idx + 1, cv_row)
                template_data.insert(r1_last_idx + 1, std_row)
                template_data.insert(r1_last_idx + 1, avg_row)
                if r2_last_idx > r1_last_idx:
                    r2_last_idx += 3

        if r2_last_idx >= 0:
            r2_rows = [r for r in template_data if str(r["编号"]) == "R2"]
            if len(r2_rows) >= 2:
                avg_row = {"参考品": "", "编号": "平均值", "质量标准": ""}
                std_row = {"参考品": "", "编号": "标准偏差", "质量标准": ""}
                cv_row = {"参考品": "", "编号": "变异系数（CV值）", "质量标准": ""}
                cv_values = {}
                for ch in channels:
                    vals = []
                    for r in r2_rows:
                        v = r.get(f"{ch}通道Ct值", "Undetermined")
                        try:
                            vals.append(float(v))
                        except (ValueError, TypeError):
                            pass
                    if len(vals) > 0:
                        a = np.mean(vals)
                        s = np.std(vals, ddof=1) if len(vals) > 1 else 0.0
                        c = round(s / a * 100, 2) if a != 0 else 0
                    else:
                        a, s, c = "/", "/", "/"
                    avg_row[f"{ch}通道Ct值"] = a if not isinstance(a, str) else a
                    std_row[f"{ch}通道Ct值"] = s if not isinstance(s, str) else s
                    cv_row[f"{ch}通道Ct值"] = f"{c}%" if not isinstance(c, str) else c
                    cv_values[ch] = c
                avg_row["检测结果"] = "/"
                avg_row["结果判读"] = "/"
                avg_row["结果判读规则"] = ""
                std_row["检测结果"] = "/"
                std_row["结果判读"] = "/"
                std_row["结果判读规则"] = ""
                cv_row["检测结果"] = "/"
                cv_ok = all(isinstance(cv_values.get(ch), (int, float)) and cv_values.get(ch) <= 5 for ch in channels if ch != "CY5")
                cv_row["结果判读"] = "符合规定" if cv_ok else ""
                cv_row["结果判读规则"] = '数值小于等于"5"'
                template_data.insert(r2_last_idx + 1, cv_row)
                template_data.insert(r2_last_idx + 1, std_row)
                template_data.insert(r2_last_idx + 1, avg_row)

        final_columns = ["参考品", "编号", "质量标准"]
        for ch in channels:
            final_columns.append(config["channel_labels"].get(ch, f"{ch}通道Ct值"))
        final_columns += ["检测结果", "结果判读", "结果判读规则"]

        df_template = pd.DataFrame(template_data)
        rename_map = {}
        for ch in channels:
            rename_map[f"{ch}通道Ct值"] = config["channel_labels"].get(ch, f"{ch}通道Ct值")
        df_template = df_template.rename(columns=rename_map)
        existing_cols = [c for c in final_columns if c in df_template.columns]
        df_template = df_template[existing_cols]

        st.subheader("📋 模板一预览")
        st.dataframe(df_template, use_container_width=True)

        output = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "原始记录附页"

        # 样式定义
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        header_font = Font(bold=True, size=10)
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        data_font = Font(size=10)
        result_col_idx = existing_cols.index("结果判读") + 1 if "结果判读" in existing_cols else None
        rule_col_idx = existing_cols.index("结果判读规则") + 1 if "结果判读规则" in existing_cols else None

        # ==================== Excel 顶部信息行 ====================
        # 第1行：日期（占参考品+编号2列，后面全部合并）
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
        ws.cell(row=1, column=1, value="日期").font = Font(bold=True, size=10)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
        if len(existing_cols) > 2:
            ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=len(existing_cols))
        for c in range(1, len(existing_cols)+1):
            ws.cell(row=1, column=c).border = thin_border

        # 第2行：企业参考品批号
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)
        ws.cell(row=2, column=1, value="企业参考品批号").font = Font(bold=True, size=10)
        ws.cell(row=2, column=1).alignment = Alignment(horizontal='center', vertical='center')
        if len(existing_cols) > 2:
            ws.merge_cells(start_row=2, start_column=3, end_row=2, end_column=len(existing_cols))
        for c in range(1, len(existing_cols)+1):
            ws.cell(row=2, column=c).border = thin_border

        # 第3行：成品批号
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=2)
        ws.cell(row=3, column=1, value="成品批号").font = Font(bold=True, size=10)
        ws.cell(row=3, column=1).alignment = Alignment(horizontal='center', vertical='center')
        if len(existing_cols) > 2:
            ws.merge_cells(start_row=3, start_column=3, end_row=3, end_column=len(existing_cols))
        for c in range(1, len(existing_cols)+1):
            ws.cell(row=3, column=c).border = thin_border

        # 第4行：规格
        ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=2)
        ws.cell(row=4, column=1, value="规格").font = Font(bold=True, size=10)
        ws.cell(row=4, column=1).alignment = Alignment(horizontal='center', vertical='center')
        if len(existing_cols) > 2:
            ws.merge_cells(start_row=4, start_column=3, end_row=4, end_column=len(existing_cols))
        for c in range(1, len(existing_cols)+1):
            ws.cell(row=4, column=c).border = thin_border

        # 第5行：表头
        header_row_num = 5
        for j, col_name in enumerate(existing_cols):
            cell = ws.cell(row=header_row_num, column=j+1, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

        # 数据行从第6行开始
        data_start_row = 6

        for i, row_data in enumerate(template_data):
            row_num = data_start_row + i
            for j, col_name in enumerate(existing_cols):
                orig_key = col_name
                for ch in channels:
                    if config["channel_labels"].get(ch) == col_name:
                        orig_key = f"{ch}通道Ct值"
                        break
                value = row_data.get(orig_key, row_data.get(col_name, ""))
                cell = ws.cell(row=row_num, column=j+1, value=value)
                cell.font = data_font
                cell.border = thin_border
                if j+1 == 3 or (rule_col_idx and j+1 == rule_col_idx):
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                if result_col_idx and j+1 == result_col_idx and value != "":
                    cell.fill = red_fill

        # 最后一行数据之后空一行，加签名行
        last_data_row = data_start_row + len(template_data) - 1
        sign_row = last_data_row + 2  # 空一行

        # 签名行：整行合并
        sign_cell = ws.cell(row=sign_row, column=1, value="检验人/日期：                                                            复核人/日期：")
        sign_cell.font = Font(size=10)
        sign_cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.merge_cells(start_row=sign_row, start_column=1, end_row=sign_row, end_column=len(existing_cols))

        # 合并参考品列、质量标准列、结果判读规则列
        merge_ranges_col1 = []
        merge_ranges_col3 = []
        merge_ranges_rule = []
        start_row = data_start_row
        prev_grp = None

        for i, row_data in enumerate(template_data):
            sample = str(row_data.get("编号", ""))
            cat = row_data.get("参考品", "")

            # 判断是否新分组：参考品列有文字，或者是统计行开始，或者R前缀变化
            new_group = False
            if cat != "" and cat is not None:
                new_group = True
            elif sample in ["平均值", "标准偏差", "变异系数（CV值）"]:
                new_group = True
            elif sample.startswith("R") and prev_grp and not prev_grp.startswith(sample[:2]):
                new_group = True

            if new_group:
                if prev_grp is not None and start_row < data_start_row + i - 1:
                    merge_ranges_col1.append((start_row, data_start_row + i - 1))
                    merge_ranges_col3.append((start_row, data_start_row + i - 1))
                    merge_ranges_rule.append((start_row, data_start_row + i - 1))
                start_row = data_start_row + i

            if sample.startswith("R"):
                prev_grp = sample[:2]
            elif sample in ["平均值", "标准偏差", "变异系数（CV值）"]:
                prev_grp = "stat"
            elif cat != "":
                prev_grp = cat

        if prev_grp is not None and start_row < data_start_row + len(template_data) - 1:
            merge_ranges_col1.append((start_row, data_start_row + len(template_data) - 1))
            merge_ranges_col3.append((start_row, data_start_row + len(template_data) - 1))
            merge_ranges_rule.append((start_row, data_start_row + len(template_data) - 1))

        for start, end in merge_ranges_col1:
            if end > start:
                ws.merge_cells(start_row=start, start_column=1, end_row=end, end_column=1)
        for start, end in merge_ranges_col3:
            if end > start:
                ws.merge_cells(start_row=start, start_column=3, end_row=end, end_column=3)
        if rule_col_idx:
            for start, end in merge_ranges_rule:
                if end > start:
                    ws.merge_cells(start_row=start, start_column=rule_col_idx, end_row=end, end_column=rule_col_idx)

        # 统计行：横向合并编号(第2列)和质量标准(第3列)
        for i, row in enumerate(template_data):
            if str(row["编号"]) in ["平均值", "标准偏差", "变异系数（CV值）"]:
                r = data_start_row + i
                ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
                ws.cell(row=r, column=2).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        wb.save(output)
        output.seek(0)

        st.download_button(
            label="📥 下载模板一 (Excel)",
            data=output,
            file_name=f"模板一_{batch_no}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        if st.button("💾 保存到历史记录"):
            record = {
                "upload_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "product_name": product_name,
                "batch_no": batch_no,
                "spec": spec,
                "inspector": inspector,
                "inspection_date": str(inspection_date),
                "data": {"template_data": template_data, "channels": channels, "project": project_name}
            }
            save_record(record)
            st.success("✅ 已保存到历史记录！")
            st.rerun()

# ==================== 历史记录 ====================
with tab2:
    st.subheader("📂 历史记录")
    records = load_all_records()

    if len(records) == 0:
        st.info("暂无历史记录")
    else:
        selected_ids = []
        for rec in records:
            col1, col2, col3, col4 = st.columns([3, 2, 2, 1])
            with col1:
                st.write(f"**{rec['product_name']}**")
            with col2:
                st.write(f"批号: {rec['batch_no']}")
            with col3:
                st.write(f"上传: {rec['upload_time']}")
            with col4:
                selected = st.checkbox("选择", key=f"sel_{rec['id']}")
                if selected:
                    selected_ids.append(rec['id'])

        st.divider()
        col_btn1, col_btn2 = st.columns(2)
        with col_btn1:
            if st.button("📥 生成模板二（合并选中记录）", disabled=len(selected_ids) < 1):
                st.info("模板二功能待开发")
        with col_btn2:
            if st.button("🗑 删除选中记录", disabled=len(selected_ids) < 1):
                delete_records(selected_ids)
                st.success("已删除")
                st.rerun()
